import openpyxl
import enchant
import string
import argparse


arg_parser = argparse.ArgumentParser(
    description="Right unknown words to .xlsx file.")
# arg_parser.add_argument('--help', type=str, required=False)
# arg_parser.add_help()


# Define input / output files, list of obvious words and class verifying words
d = enchant.Dict("en")
xlsx_file = 'words.xlsx'  # Known words
subs = "sub.srt"
output_file = 'new_words.xlsx'
exlude_list = ['a', 'an', 'the']  # List of obvious words, for manual appending


# Fill the list of words we know.
words_i_know = []
w = openpyxl.load_workbook(xlsx_file).active
for row in w:
    for cell in row:
        if cell.value:
            words_i_know.append(cell.value)

# Read contents of the sub file.
with open(subs) as s:
    content = s.readlines()

# Fill the list of known words, removing non-letter characters and non-english words.
words_in_subs = []
for line in content:
    for word in line.split(' '):
        for i, letter in enumerate(word):
            if letter not in string.ascii_letters and letter != "'":
                word = word.replace(letter, '')
        if word != '':
            words_in_subs.append(word.lower())


# Copy of the words in subs for later print out.
all_words_in_subs = list(set(words_in_subs))

words_in_subs = set(words_in_subs) - set(exlude_list)  # exlude obvious words.

# Exlude known words.
words_in_subs = sorted(list(words_in_subs))
for word in words_i_know:
    word = str(word)
    if word.lower() in words_in_subs:
        words_in_subs.remove(word.lower())

# Save results to .xlsx file.
wb = openpyxl.Workbook()
ws = wb.active
for i, word in enumerate(words_in_subs, start=1):
    ws.cell(row=i, column=1, value=word)
wb.save('new_words.xlsx')

# Print out of the results.
print("Known words: " + str(words_i_know), end="\n\n")
print("Words in subs: " + str(all_words_in_subs), end="\n\n")
print("Words in subs minus known words: " + str(words_in_subs), end="\n\n")

